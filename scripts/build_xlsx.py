#!/usr/bin/env python3
"""
Build Conditioning_Atlas.xlsx from merged.json + excluded_*.tsv.

Sheets produced:
    README         — what this is, how to read it, provenance
    Atlas          — the data, one row per experiment, filterable
    Coverage gaps  — live COUNTIFS formulas over the Atlas sheet + narrative gap list
    Data dictionary— column definitions
    Screened out   — papers that were checked and rejected, with reason

Run this after scripts/merge_rows.py any time merged.json changes.

Usage:
    python3 scripts/build_xlsx.py
    python3 scripts/recalc_xlsx.py Conditioning_Atlas.xlsx   # then recalculate formulas (see that script)

Requires: openpyxl (pip install openpyxl --break-system-packages if missing).
"""
import datetime
import glob
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FONT = "Arial"
NAVY = "1F3864"

HDR = [
    "study_id", "doi", "pmid", "year", "species", "taxon", "stressor", "stressor_category",
    "conditioning_stage", "dose_treatment", "duration", "generations", "transmission", "challenge",
    "persistence_days", "multiple_timepoints", "decay", "readouts", "direction", "effect_magnitude",
    "molecular_layers", "concordance", "applied_framing", "replication", "evidence_source",
    "editorial_notice", "notes", "cluster",
]
PRETTY = {
    "study_id": "Study ID", "doi": "DOI", "pmid": "PMID", "year": "Year", "species": "Species",
    "taxon": "Taxon", "stressor": "Stressor", "stressor_category": "Stressor category",
    "conditioning_stage": "Conditioning stage", "dose_treatment": "Dose (treatment vs control)",
    "duration": "Conditioning duration", "generations": "Generations", "transmission": "Transmission route",
    "challenge": "Later challenge", "persistence_days": "Persistence (days)",
    "multiple_timepoints": "Multiple timepoints?", "decay": "Decay", "readouts": "Phenotypic readouts",
    "direction": "Direction of effect", "effect_magnitude": "Effect magnitude",
    "molecular_layers": "Molecular layers", "concordance": "Molecular-phenotype concordance",
    "applied_framing": "Applied framing", "replication": "Replication", "evidence_source": "Evidence source",
    "editorial_notice": "Editorial notice", "notes": "Notes", "cluster": "Cluster",
}
WIDTHS = {
    "study_id": 16, "doi": 30, "pmid": 10, "year": 7, "species": 26, "taxon": 15, "stressor": 26,
    "stressor_category": 15, "conditioning_stage": 16, "dose_treatment": 42, "duration": 34,
    "generations": 11, "transmission": 16, "challenge": 34, "persistence_days": 13,
    "multiple_timepoints": 12, "decay": 11, "readouts": 40, "direction": 13, "effect_magnitude": 46,
    "molecular_layers": 30, "concordance": 34, "applied_framing": 14, "replication": 30,
    "evidence_source": 14, "editorial_notice": 14, "notes": 60, "cluster": 30,
}

DEFS = [
    ("Study ID", "First author surname + year; a/b suffix distinguishes separate experiments from one paper.", "Lastname_Year"),
    ("DOI", "Bare DOI. Prefix with https://doi.org/ to open.", "text"),
    ("PMID", "PubMed ID where indexed.", "integer or 'not reported'"),
    ("Year", "Publication year.", "integer"),
    ("Species", "Full Latin binomial(s) of the conditioned organism.", "text"),
    ("Taxon", "Broad taxonomic group.", "Bivalve | Coral/cnidarian | Gastropod | Echinoderm | Crustacean | Polychaete/other | Multiple"),
    ("Stressor", "The specific stressor used in the CONDITIONING exposure.", "text"),
    ("Stressor category", "Binned stressor for filtering.", "Temperature | pH/OA | Hypoxia | Salinity | Pollutant | Pathogen/immune | Multiple"),
    ("Conditioning stage", "Life stage the priming exposure was applied to — the 'programming window'.", "Broodstock/adult | Gamete | Embryo | Larva | Juvenile | Multiple"),
    ("Dose (treatment vs control)", "Quantitative intensity of the conditioning exposure, always paired with the control value.", "text with units"),
    ("Conditioning duration", "How long the conditioning exposure lasted; cycling regimes described.", "text with units"),
    ("Generations", "Which generations were measured, using the paper's own labels.", "F0 | F0-F1 | F0-F1-F2 | ..."),
    ("Transmission route", "How the effect reached the assessed animals.", "Maternal | Paternal | Both parents | Route not resolved | Within-generation only"),
    ("Later challenge", "The stress the conditioned animals or their offspring were tested against.", "text; 'none (baseline only)' if unchallenged"),
    ("Persistence (days)", "Days from the END of conditioning to the assessment or challenge. 0 = offspring spawned straight into the challenge. '~' marks a value derived from the paper's own numbers. CAVEAT: a small number of rows were conditioned continuously or repeatedly rather than conditioned-then-released, so their value is an exposure/tracking window, not a post-conditioning recovery gap. Those rows say so explicitly in Notes. Do not rank them against clean-gap rows.", "number, '~number', or 'not reported'"),
    ("Multiple timepoints?", "Did the study assess at more than one post-conditioning time?", "Yes | No"),
    ("Decay", "What happened to the induced state over time.", "Persisted | Decayed | Strengthened | Not tested"),
    ("Phenotypic readouts", "Organismal and physiological outcomes measured.", "text, semicolon-separated"),
    ("Direction of effect", "Effect of conditioning on performance under the later challenge, relative to naive controls.", "Improved | No effect | Worsened | Mixed | Not assessable"),
    ("Effect magnitude", "Quoted numbers with the statistic as reported by the authors.", "text"),
    ("Molecular layers", "Which regulatory layers were assayed, with method.", "Methylation (WGBS/methylRAD/MBD-BS) | Transcriptome (RNA-seq) | Targeted qPCR | ncRNA | Chromatin | None"),
    ("Molecular-phenotype concordance", "Whether molecular change tracked the phenotypic outcome, with any quantitative overlap.", "Concordant | Discordant | Partial | Not evaluated"),
    ("Applied framing", "Whether the paper itself frames the work as applied.", "Aquaculture | Restoration | Both | Biomonitoring | None"),
    ("Replication", "Tanks/biological replicates per treatment, families, individuals.", "text"),
    ("Evidence source", "How deeply the row was sourced.", "Abstract only | Full text | Abstract + citations"),
    ("Editorial notice", "Scite editorial-notice check result at time of extraction.", "None found | Retracted | Correction | Concern | Not checked"),
    ("Notes", "Caveats, trade-offs, maladaptive findings, and why a field is unreported.", "text"),
    ("Cluster", "Which extraction sweep the row came from. Retained for provenance.", "text"),
]

TAXA = ["Bivalve", "Coral/cnidarian", "Crustacean", "Echinoderm", "Polychaete/other", "Gastropod"]
STRESSORS = ["Temperature", "Pathogen/immune", "pH/OA", "Multiple", "Pollutant", "Hypoxia", "Salinity"]
STAGES = ["Broodstock/adult", "Juvenile", "Multiple", "Embryo", "Larva"]
DIRECTIONS = ["Improved", "Mixed", "Worsened", "No effect", "Not assessable"]
DECAY = ["Persisted", "Not tested", "Decayed", "Strengthened"]
FRAMING = ["Aquaculture", "None", "Restoration", "Biomonitoring", "Both"]


def L(name):
    return get_column_letter(HDR.index(name) + 1)


def load_rows():
    path = os.path.join(ROOT, "merged.json")
    with open(path) as fh:
        return json.load(fh)


def load_excluded():
    labels = {
        "excluded_bivalve_temp_ph.tsv": "Bivalve: temperature & pH",
        "excluded_coral.tsv": "Coral / cnidarian",
        "excluded_immune.tsv": "Immune priming",
        "excluded_other.tsv": "Hypoxia / salinity / pollutant / other taxa",
    }
    out = []
    for path in sorted(glob.glob(os.path.join(ROOT, "excluded_*.tsv"))):
        fname = os.path.basename(path)
        label = labels.get(fname, fname.replace("excluded_", "").replace(".tsv", ""))
        with open(path) as fh:
            for line in fh:
                line = line.rstrip("\n")
                if not line.strip():
                    continue
                parts = line.split("\t")
                while len(parts) < 3:
                    parts.append("")
                out.append((parts[0], parts[1], parts[2], label))
    return out


def build(rows, excluded, out_path):
    wb = Workbook()
    HEADFILL = PatternFill("solid", fgColor=NAVY)
    thin = Side(style="thin", color="D9D9D9")
    BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
    GREEN = PatternFill("solid", fgColor="E2EFDA")
    RED = PatternFill("solid", fgColor="FCE4E4")
    AMBER = PatternFill("solid", fgColor="FFF2CC")
    GREY = PatternFill("solid", fgColor="F2F2F2")

    # ---------- Atlas ----------
    ws = wb.active
    ws.title = "Atlas"
    ws.append([PRETTY[h] for h in HDR])
    for c in ws[1]:
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = HEADFILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORD
    for r in rows:
        ws.append([r.get(h, "") for h in HDR])
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HDR))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 34
    for i, h in enumerate(HDR, 1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[h]
    dircol = HDR.index("direction") + 1
    decaycol = HDR.index("decay") + 1
    for ri in range(2, len(rows) + 2):
        for ci in range(1, len(HDR) + 1):
            c = ws.cell(row=ri, column=ci)
            c.font = Font(name=FONT, size=9)
            c.border = BORD
            c.alignment = Alignment(vertical="top", wrap_text=True)
        d = ws.cell(row=ri, column=dircol).value
        ws.cell(row=ri, column=dircol).fill = {"Improved": GREEN, "Worsened": RED, "Mixed": AMBER}.get(d, GREY)
        dk = ws.cell(row=ri, column=decaycol).value
        ws.cell(row=ri, column=decaycol).fill = {"Persisted": GREEN, "Decayed": RED, "Not tested": GREY, "Strengthened": GREEN}.get(dk, GREY)
        ws.row_dimensions[ri].height = 58

    # ---------- Coverage gaps ----------
    cg = wb.create_sheet("Coverage gaps")
    N = len(rows) + 1

    def head(ws_, row, span=3):
        for k in range(1, span + 1):
            ws_.cell(row=row, column=k).fill = HEADFILL

    cg["A1"] = "Conditioning Atlas — coverage and gaps"
    cg["A1"].font = Font(name=FONT, bold=True, size=14, color=NAVY)
    cg["A2"] = f"All counts are live COUNTIFS/COUNTA formulas over Atlas!A2:AB{N}. Edit or add rows in Atlas and these update."
    cg["A2"].font = Font(name=FONT, italic=True, size=9, color="808080")
    cg["B1"] = f"=COUNTA(Atlas!A2:A{N})"
    cg["B1"].font = Font(name=FONT, size=9, color="808080")
    cg["C1"] = "rows in Atlas"
    cg["C1"].font = Font(name=FONT, size=9, color="808080")

    def block(startrow, title, col_letter, values):
        cg.cell(row=startrow, column=1, value=title).font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
        head(cg, startrow)
        cg.cell(row=startrow + 1, column=1, value="Category").font = Font(name=FONT, bold=True, size=10)
        cg.cell(row=startrow + 1, column=2, value="Studies").font = Font(name=FONT, bold=True, size=10)
        cg.cell(row=startrow + 1, column=3, value="% of atlas").font = Font(name=FONT, bold=True, size=10)
        r = startrow + 2
        for v in values:
            cg.cell(row=r, column=1, value=v).font = Font(name=FONT, size=10)
            cg.cell(row=r, column=2, value=f'=COUNTIFS(Atlas!{col_letter}$2:{col_letter}${N},A{r})').font = Font(name=FONT, size=10)
            c = cg.cell(row=r, column=3, value=f'=IFERROR(B{r}/$B${startrow + 2 + len(values)},0)')
            c.font = Font(name=FONT, size=10)
            c.number_format = "0.0%"
            r += 1
        cg.cell(row=r, column=1, value="TOTAL").font = Font(name=FONT, bold=True, size=10)
        cg.cell(row=r, column=2, value=f'=SUM(B{startrow + 2}:B{r - 1})').font = Font(name=FONT, bold=True, size=10)
        return r + 2

    r = 4
    r = block(r, "Taxonomic coverage", L("taxon"), TAXA)
    r = block(r, "Stressor coverage", L("stressor_category"), STRESSORS)
    r = block(r, "Conditioning window", L("conditioning_stage"), STAGES)
    r = block(r, "Direction of effect", L("direction"), DIRECTIONS)
    r = block(r, "Was decay of the induced state tested?", L("decay"), DECAY)
    r = block(r, "Applied framing", L("applied_framing"), FRAMING)

    cg.cell(row=r, column=1, value="Priority-field completeness").font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    head(cg, r)
    cg.cell(row=r + 1, column=1, value="Field").font = Font(name=FONT, bold=True, size=10)
    cg.cell(row=r + 1, column=2, value="Rows with a value").font = Font(name=FONT, bold=True, size=10)
    cg.cell(row=r + 1, column=3, value="Rows 'not reported'").font = Font(name=FONT, bold=True, size=10)
    comp = [
        ("Persistence (days)", L("persistence_days")), ("Dose (treatment vs control)", L("dose_treatment")),
        ("Conditioning duration", L("duration")), ("Effect magnitude", L("effect_magnitude")),
        ("Molecular-phenotype concordance", L("concordance")), ("Replication", L("replication")),
    ]
    rr = r + 2
    for label, col in comp:
        cg.cell(row=rr, column=1, value=label).font = Font(name=FONT, size=10)
        cg.cell(row=rr, column=2, value=f'=$B$1-COUNTIF(Atlas!{col}$2:{col}${N},"*not reported*")').font = Font(name=FONT, size=10)
        cg.cell(row=rr, column=3, value=f'=COUNTIF(Atlas!{col}$2:{col}${N},"*not reported*")').font = Font(name=FONT, size=10)
        rr += 1
    rr += 1

    cg.cell(row=rr, column=1, value="Structural gaps this atlas exposes (update by hand as the atlas grows)").font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    head(cg, rr)
    gaps = [
        "Only a handful of rows condition at the embryo or larval stage, despite embryo-larval development being the framework's named programming window. Most rows condition broodstock or adults instead.",
        "Decay of the induced state is untested in roughly half of rows. Where it is tested, it decays about as often as it persists.",
        "No immune-priming row assays DNA methylation; the epigenetic work there is all histone marks, on a timescale that contradicts the measured protection duration.",
        "No row in the hypoxia or salinity categories pairs a molecular layer with a decay time-course.",
        "Gastropods, barnacles, crabs, and amphipods are essentially absent. Salinity as a primary stressor is barely represented.",
        "Molecular-phenotype concordance is unevaluated in most rows that measure both layers.",
        "PERSISTENCE SEMANTICS: persistence_days means different things for a conditioned-then-released design versus a continuously/repeatedly exposed one. Affected rows flag this in Notes — do not rank them against each other without reading the flag.",
    ]
    for i, g in enumerate(gaps):
        c = cg.cell(row=rr + 1 + i, column=1, value=f"• {g}")
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        cg.merge_cells(start_row=rr + 1 + i, start_column=1, end_row=rr + 1 + i, end_column=3)
        cg.row_dimensions[rr + 1 + i].height = 30
    cg.column_dimensions["A"].width = 78
    cg.column_dimensions["B"].width = 18
    cg.column_dimensions["C"].width = 14

    # ---------- Data dictionary ----------
    dd = wb.create_sheet("Data dictionary")
    dd.append(["Column", "Definition", "Allowed values / format"])
    for row in DEFS:
        dd.append(list(row))
    for c in dd[1]:
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = HEADFILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for ri in range(2, dd.max_row + 1):
        for ci in range(1, 4):
            c = dd.cell(row=ri, column=ci)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORD
        dd.row_dimensions[ri].height = 40 if ri != 15 else 96  # row 15 = Persistence, long caveat
    dd.column_dimensions["A"].width = 32
    dd.column_dimensions["B"].width = 72
    dd.column_dimensions["C"].width = 52
    dd.freeze_panes = "A2"

    # ---------- Screened out ----------
    ex = wb.create_sheet("Screened out")
    ex.append(["DOI / identifier", "Title", "Reason for exclusion", "Sweep"])
    for row in excluded:
        ex.append(list(row))
    for c in ex[1]:
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = HEADFILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for ri in range(2, ex.max_row + 1):
        for ci in range(1, 5):
            c = ex.cell(row=ri, column=ci)
            c.font = Font(name=FONT, size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORD
        ex.row_dimensions[ri].height = 28
    ex.column_dimensions["A"].width = 34
    ex.column_dimensions["B"].width = 70
    ex.column_dimensions["C"].width = 50
    ex.column_dimensions["D"].width = 32
    ex.freeze_panes = "A2"
    ex.auto_filter.ref = f"A1:D{ex.max_row}"

    # ---------- README ----------
    rd = wb.create_sheet("README", 0)
    rd["A1"] = "The Conditioning Atlas"
    rd["A1"].font = Font(name=FONT, bold=True, size=18, color=NAVY)
    rd["A2"] = "Priming windows, doses, and persistence of induced environmental memory in marine invertebrates"
    rd["A2"].font = Font(name=FONT, italic=True, size=11, color="595959")
    npap = len(set(r["doi"].lower() for r in rows if r["doi"] != "not reported"))
    notes = [
        "",
        f"{len(rows)} experiments extracted from {npap} papers. One row per experiment: a paper running two "
        "stressors or two species appears more than once, with the same DOI and an a/b suffix on the Study ID.",
        "",
        "WHAT THIS IS FOR",
        "The lab framework states that the application program for environmental conditioning is to 'identify the "
        "windows, identify the doses, and characterize how long the induced memory persists relative to commercial "
        "production timescales.' Nobody had that in tabular form. This is that table.",
        "",
        "SHEETS",
        "Atlas — the data. Filterable. Direction and Decay are colour-coded (green = beneficial/persisted, "
        "red = worsened/decayed, grey = not tested).",
        "Coverage gaps — live COUNTIFS over the Atlas sheet, plus the structural gaps the atlas exposes. Add rows "
        "to Atlas and these recalculate.",
        "Data dictionary — what every column means and its allowed values.",
        f"Screened out — the {len(excluded)} papers that were checked and rejected, with the reason. Kept so the "
        "screening is auditable and not repeated.",
        "",
        "HOW TO READ THE PRIORITY COLUMNS",
        "Persistence (days) is days from the END of conditioning to the assessment. A 0 means offspring were "
        "spawned straight into the challenge. A '~' means the value was derived from the paper's own reported "
        "numbers rather than stated outright.",
        "Decay = 'Not tested' is itself the finding: most of this literature measures one timepoint and cannot "
        "distinguish a persistent mark from a transient one.",
        "'not reported' is deliberate. No value was inferred, estimated, or filled in from a similar study.",
        "",
        "PROVENANCE AND LIMITS",
        "Assembled from PubMed, Consensus, and Scite full-text retrieval. Every DOI was resolved through Scite and "
        "checked for retractions, corrections, and expressions of concern. A verification pass then re-checked a "
        "sample of rows line-by-line against full text; see git history / CHANGELOG for what that pass corrected.",
        "This is not a registered systematic review: no PRISMA flow, no dual independent extraction, no formal "
        "risk-of-bias assessment. Treat it as a working lab resource and a scaffold for a real review.",
        "See the repository README.md for the full maintenance and extension procedure.",
        "",
        f"Built {datetime.date.today().isoformat()}.",
    ]
    r = 3
    for t in notes:
        c = rd.cell(row=r, column=1, value=t)
        bold = t.isupper() and len(t) < 40
        c.font = Font(name=FONT, size=11, bold=bold, color=NAVY if bold else "000000")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rd.row_dimensions[r].height = 15 if not t else (18 if bold else max(15, 15 * (len(t) // 95 + 1)))
        r += 1
    rd.column_dimensions["A"].width = 118

    wb.save(out_path)
    return len(rows), len(excluded)


def main():
    rows = load_rows()
    excluded = load_excluded()
    out_path = os.path.join(ROOT, "Conditioning_Atlas.xlsx")
    n_rows, n_ex = build(rows, excluded, out_path)
    print(f"Saved {out_path} — {n_rows} atlas rows, {n_ex} screened-out entries.")
    print("Next: python3 scripts/recalc_xlsx.py Conditioning_Atlas.xlsx   (to compute the Coverage gaps formulas)")


if __name__ == "__main__":
    main()
