#!/usr/bin/env python3
"""Build data/atlas.xlsx from data/atlas.csv + the schema.

The .xlsx is a convenience artifact for people who live in spreadsheets. It adds:
  - an AutoFilter on every column
  - a frozen header row
  - data-validation dropdowns on enum columns (so hand-edits stay in-vocabulary)
  - a second sheet, "flagship", pre-filtered to bivalve rows where persistence past
    metamorphosis was actually tested
  - a "schema" sheet documenting each column

data/atlas.csv remains the source of truth; regenerate the xlsx after edits.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import yaml
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
VOCAB = ROOT / "data" / "schema" / "controlled_vocab.yaml"
ATLAS = ROOT / "data" / "atlas.csv"
OUT = ROOT / "data" / "atlas.xlsx"


def _norm(v):
    if v is True:
        return "yes"
    if v is False:
        return "no"
    return str(v)


def load_spec() -> list[dict]:
    with open(VOCAB) as fh:
        cols = yaml.safe_load(fh)["columns"]
    for c in cols:
        if "values" in c:
            c["values"] = [_norm(v) for v in c["values"]]
    return cols


def main() -> int:
    spec = load_spec()
    columns = [c["name"] for c in spec]
    by_name = {c["name"]: c for c in spec}

    df = pd.read_csv(ATLAS, dtype=str).fillna("")
    df = df[columns]  # enforce schema order

    flagship = df[(df["taxon_group"] == "bivalve")
                  & (df["persisted_past_metamorphosis"].isin(["yes", "no"]))]

    schema_df = pd.DataFrame([
        {
            "column": c["name"],
            "type": c["type"],
            "required": bool(c.get("required")),
            "allowed_values": ", ".join(c["values"]) if "values" in c else "",
            "note": c.get("note", ""),
        }
        for c in spec
    ])

    with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name="atlas", index=False)
        flagship.to_excel(xl, sheet_name="flagship", index=False)
        schema_df.to_excel(xl, sheet_name="schema", index=False)

        wb = xl.book
        ws = wb["atlas"]
        n_rows = len(df) + 1
        n_cols = len(columns)
        last_col = get_column_letter(n_cols)

        ws.auto_filter.ref = f"A1:{last_col}{n_rows}"
        ws.freeze_panes = "A2"

        # dropdowns on enum columns, applied to a generous row range for future edits
        max_edit_row = max(n_rows, 2000)
        for idx, name in enumerate(columns, start=1):
            c = by_name[name]
            if c["type"] != "enum":
                continue
            allowed = list(c["values"]) + ["NA", "NR"]
            formula = '"' + ",".join(allowed) + '"'
            if len(formula) > 255:
                continue  # Excel inline-list limit; skip dropdown, validator still enforces
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            col_letter = get_column_letter(idx)
            dv.add(f"{col_letter}2:{col_letter}{max_edit_row}")
            ws.add_data_validation(dv)

        # reasonable column widths
        for idx, name in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(max(len(name) + 2, 12), 42)

        for sheet in ("flagship", "schema"):
            s = wb[sheet]
            s.freeze_panes = "A2"
            ncol = s.max_column
            s.auto_filter.ref = f"A1:{get_column_letter(ncol)}{s.max_row}"

    print(f"wrote {OUT}  (atlas={len(df)} rows, flagship={len(flagship)} rows)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
